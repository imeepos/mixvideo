#!/usr/bin/env python3
"""
媒体资源扫描工具

扫描指定目录及子目录下的所有视频/音频/图片等资源，生成资源清单JSON文件
"""

import os
import json
import hashlib
from pathlib import Path
from typing import Dict, List, Optional, Union, Any
from dataclasses import dataclass, asdict
import mimetypes
import time
from datetime import datetime


@dataclass
class MediaInfo:
    """媒体文件信息"""
    file_path: str
    file_name: str
    file_size: int
    file_extension: str
    mime_type: str
    media_type: str  # video, audio, image
    created_time: str
    modified_time: str
    file_hash: Optional[str] = None
    duration: Optional[float] = None  # 视频/音频时长（秒）
    width: Optional[int] = None      # 视频/图片宽度
    height: Optional[int] = None     # 视频/图片高度
    bitrate: Optional[int] = None    # 比特率
    fps: Optional[float] = None      # 帧率（视频）


class MediaScanner:
    """媒体资源扫描器"""
    
    # 支持的媒体格式
    VIDEO_EXTENSIONS = {
        '.mp4', '.avi', '.mov', '.mkv', '.wmv', '.flv', '.webm', 
        '.m4v', '.3gp', '.ts', '.mts', '.m2ts', '.vob', '.asf'
    }
    
    AUDIO_EXTENSIONS = {
        '.mp3', '.wav', '.flac', '.aac', '.ogg', '.wma', '.m4a',
        '.opus', '.aiff', '.au', '.ra', '.amr', '.ac3'
    }
    
    IMAGE_EXTENSIONS = {
        '.jpg', '.jpeg', '.png', '.gif', '.bmp', '.tiff', '.tif',
        '.webp', '.svg', '.ico', '.psd', '.raw', '.cr2', '.nef'
    }
    
    def __init__(self, include_hash: bool = False, include_metadata: bool = True):
        """
        初始化扫描器
        
        Args:
            include_hash: 是否计算文件哈希值
            include_metadata: 是否提取媒体元数据
        """
        self.include_hash = include_hash
        self.include_metadata = include_metadata
        
        # 初始化mimetypes
        mimetypes.init()
    
    def get_media_type(self, file_path: Path) -> Optional[str]:
        """
        根据文件扩展名判断媒体类型
        
        Args:
            file_path: 文件路径
            
        Returns:
            媒体类型: 'video', 'audio', 'image' 或 None
        """
        ext = file_path.suffix.lower()
        
        if ext in self.VIDEO_EXTENSIONS:
            return 'video'
        elif ext in self.AUDIO_EXTENSIONS:
            return 'audio'
        elif ext in self.IMAGE_EXTENSIONS:
            return 'image'
        else:
            return None
    
    def calculate_file_hash(self, file_path: Path) -> str:
        """
        计算文件MD5哈希值
        
        Args:
            file_path: 文件路径
            
        Returns:
            MD5哈希值
        """
        hash_md5 = hashlib.md5()
        try:
            with open(file_path, "rb") as f:
                for chunk in iter(lambda: f.read(4096), b""):
                    hash_md5.update(chunk)
            return hash_md5.hexdigest()
        except Exception:
            return None
    
    def get_media_metadata(self, file_path: Path, media_type: str) -> Dict[str, Any]:
        """
        提取媒体元数据
        
        Args:
            file_path: 文件路径
            media_type: 媒体类型
            
        Returns:
            元数据字典
        """
        metadata = {}
        
        try:
            if media_type in ['video', 'audio']:
                # 尝试使用ffprobe获取视频/音频信息
                metadata.update(self._get_ffprobe_info(file_path))
            elif media_type == 'image':
                # 尝试使用PIL获取图片信息
                metadata.update(self._get_image_info(file_path))
        except Exception as e:
            print(f"获取元数据失败 {file_path}: {e}")
        
        return metadata
    
    def _get_ffprobe_info(self, file_path: Path) -> Dict[str, Any]:
        """使用ffprobe获取视频/音频信息"""
        import subprocess
        import json as json_module
        
        try:
            cmd = [
                'ffprobe',
                '-v', 'quiet',
                '-print_format', 'json',
                '-show_format',
                '-show_streams',
                str(file_path)
            ]
            
            result = subprocess.run(cmd, capture_output=True, text=True, timeout=30)
            
            if result.returncode == 0:
                data = json_module.loads(result.stdout)
                
                metadata = {}
                
                # 获取格式信息
                if 'format' in data:
                    format_info = data['format']
                    if 'duration' in format_info:
                        metadata['duration'] = float(format_info['duration'])
                    if 'bit_rate' in format_info:
                        metadata['bitrate'] = int(format_info['bit_rate'])
                
                # 获取流信息
                if 'streams' in data:
                    for stream in data['streams']:
                        if stream.get('codec_type') == 'video':
                            if 'width' in stream:
                                metadata['width'] = stream['width']
                            if 'height' in stream:
                                metadata['height'] = stream['height']
                            if 'r_frame_rate' in stream:
                                fps_str = stream['r_frame_rate']
                                if '/' in fps_str:
                                    num, den = fps_str.split('/')
                                    if int(den) != 0:
                                        metadata['fps'] = float(num) / float(den)
                
                return metadata
        except Exception:
            pass
        
        return {}
    
    def _get_image_info(self, file_path: Path) -> Dict[str, Any]:
        """使用PIL获取图片信息"""
        try:
            from PIL import Image
            
            with Image.open(file_path) as img:
                return {
                    'width': img.width,
                    'height': img.height,
                    'format': img.format
                }
        except Exception:
            pass
        
        return {}
    
    def scan_file(self, file_path: Path) -> Optional[MediaInfo]:
        """
        扫描单个文件
        
        Args:
            file_path: 文件路径
            
        Returns:
            媒体信息对象或None
        """
        # 检查是否为媒体文件
        media_type = self.get_media_type(file_path)
        if not media_type:
            return None
        
        try:
            # 获取文件基本信息
            stat = file_path.stat()
            
            # 获取MIME类型
            mime_type, _ = mimetypes.guess_type(str(file_path))
            if not mime_type:
                mime_type = 'application/octet-stream'
            
            # 创建媒体信息对象
            media_info = MediaInfo(
                file_path=str(file_path.resolve()),
                file_name=file_path.name,
                file_size=stat.st_size,
                file_extension=file_path.suffix.lower(),
                mime_type=mime_type,
                media_type=media_type,
                created_time=datetime.fromtimestamp(stat.st_ctime).isoformat(),
                modified_time=datetime.fromtimestamp(stat.st_mtime).isoformat()
            )
            
            # 计算文件哈希值
            if self.include_hash:
                media_info.file_hash = self.calculate_file_hash(file_path)
            
            # 提取媒体元数据
            if self.include_metadata:
                metadata = self.get_media_metadata(file_path, media_type)
                if metadata:
                    media_info.duration = metadata.get('duration')
                    media_info.width = metadata.get('width')
                    media_info.height = metadata.get('height')
                    media_info.bitrate = metadata.get('bitrate')
                    media_info.fps = metadata.get('fps')
            
            return media_info
            
        except Exception as e:
            print(f"扫描文件失败 {file_path}: {e}")
            return None
    
    def scan_directory(self, directory: Union[str, Path], 
                      recursive: bool = True,
                      progress_callback: Optional[callable] = None) -> List[MediaInfo]:
        """
        扫描目录
        
        Args:
            directory: 目录路径
            recursive: 是否递归扫描子目录
            progress_callback: 进度回调函数
            
        Returns:
            媒体信息列表
        """
        directory = Path(directory)
        
        if not directory.exists() or not directory.is_dir():
            raise ValueError(f"目录不存在或不是有效目录: {directory}")
        
        media_files = []
        
        # 获取所有文件
        if recursive:
            pattern = "**/*"
        else:
            pattern = "*"
        
        all_files = list(directory.glob(pattern))
        total_files = len(all_files)
        
        print(f"开始扫描目录: {directory}")
        print(f"发现 {total_files} 个文件")
        
        processed = 0
        for file_path in all_files:
            if file_path.is_file():
                media_info = self.scan_file(file_path)
                if media_info:
                    media_files.append(media_info)
                
                processed += 1
                
                # 调用进度回调
                if progress_callback:
                    progress_callback(processed, total_files, file_path)
                
                # 简单进度显示
                if processed % 100 == 0 or processed == total_files:
                    print(f"已处理: {processed}/{total_files} 文件，发现媒体文件: {len(media_files)}")
        
        print(f"扫描完成，共发现 {len(media_files)} 个媒体文件")
        return media_files
    
    def generate_inventory(self, media_files: List[MediaInfo]) -> Dict[str, Any]:
        """
        生成资源清单
        
        Args:
            media_files: 媒体文件列表
            
        Returns:
            资源清单字典
        """
        # 按类型分组
        by_type = {'video': [], 'audio': [], 'image': []}
        total_size = 0
        
        for media in media_files:
            by_type[media.media_type].append(asdict(media))
            total_size += media.file_size
        
        # 统计信息
        stats = {
            'total_files': len(media_files),
            'total_size': total_size,
            'total_size_mb': round(total_size / (1024 * 1024), 2),
            'video_count': len(by_type['video']),
            'audio_count': len(by_type['audio']),
            'image_count': len(by_type['image']),
            'scan_time': datetime.now().isoformat()
        }
        
        # 扩展名统计
        extensions = {}
        for media in media_files:
            ext = media.file_extension
            if ext not in extensions:
                extensions[ext] = {'count': 0, 'size': 0}
            extensions[ext]['count'] += 1
            extensions[ext]['size'] += media.file_size
        
        return {
            'metadata': {
                'scan_time': datetime.now().isoformat(),
                'scanner_version': '1.0.0',
                'include_hash': self.include_hash,
                'include_metadata': self.include_metadata
            },
            'statistics': stats,
            'extensions': extensions,
            'files': {
                'videos': by_type['video'],
                'audios': by_type['audio'],
                'images': by_type['image']
            }
        }
    
    def save_inventory(self, inventory: Dict[str, Any], output_path: Union[str, Path],
                      format_type: str = 'json'):
        """
        保存资源清单到文件

        Args:
            inventory: 资源清单
            output_path: 输出文件路径
            format_type: 输出格式 ('json', 'csv', 'html', 'markdown', 'excel')
        """
        output_path = Path(output_path)

        # 确保输出目录存在
        output_path.parent.mkdir(parents=True, exist_ok=True)

        if format_type.lower() == 'json':
            self._save_json(inventory, output_path)
        elif format_type.lower() == 'csv':
            self._save_csv(inventory, output_path)
        elif format_type.lower() == 'html':
            self._save_html(inventory, output_path)
        elif format_type.lower() == 'markdown':
            self._save_markdown(inventory, output_path)
        elif format_type.lower() == 'excel':
            self._save_excel(inventory, output_path)
        else:
            raise ValueError(f"不支持的格式类型: {format_type}")

        print(f"资源清单已保存到: {output_path}")

    def _save_json(self, inventory: Dict[str, Any], output_path: Path):
        """保存为JSON格式"""
        with open(output_path, 'w', encoding='utf-8') as f:
            json.dump(inventory, f, indent=2, ensure_ascii=False)

    def _save_csv(self, inventory: Dict[str, Any], output_path: Path):
        """保存为CSV格式"""
        import csv

        # 准备CSV数据
        csv_data = []

        # 添加所有媒体文件
        for media_type in ['videos', 'audios', 'images']:
            for file_info in inventory['files'][media_type]:
                row = {
                    '文件名': file_info['file_name'],
                    '文件路径': file_info['file_path'],
                    '媒体类型': file_info['media_type'],
                    '文件大小(字节)': file_info['file_size'],
                    '文件大小(MB)': round(file_info['file_size'] / (1024 * 1024), 2),
                    '扩展名': file_info['file_extension'],
                    'MIME类型': file_info['mime_type'],
                    '创建时间': file_info['created_time'],
                    '修改时间': file_info['modified_time']
                }

                # 添加可选字段
                if file_info.get('file_hash'):
                    row['文件哈希'] = file_info['file_hash']
                if file_info.get('duration'):
                    row['时长(秒)'] = file_info['duration']
                if file_info.get('width'):
                    row['宽度'] = file_info['width']
                if file_info.get('height'):
                    row['高度'] = file_info['height']
                if file_info.get('fps'):
                    row['帧率'] = file_info['fps']
                if file_info.get('bitrate'):
                    row['比特率'] = file_info['bitrate']

                csv_data.append(row)

        # 写入CSV文件
        if csv_data:
            fieldnames = csv_data[0].keys()
            with open(output_path, 'w', newline='', encoding='utf-8') as f:
                writer = csv.DictWriter(f, fieldnames=fieldnames)
                writer.writeheader()
                writer.writerows(csv_data)

    def _save_html(self, inventory: Dict[str, Any], output_path: Path):
        """保存为HTML表格格式"""
        stats = inventory['statistics']

        html_content = f"""<!DOCTYPE html>
<html lang="zh-CN">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>媒体资源清单</title>
    <style>
        body {{ font-family: Arial, sans-serif; margin: 20px; }}
        .header {{ background-color: #f5f5f5; padding: 20px; border-radius: 5px; margin-bottom: 20px; }}
        .stats {{ display: flex; gap: 20px; margin-bottom: 20px; }}
        .stat-card {{ background-color: #e3f2fd; padding: 15px; border-radius: 5px; text-align: center; }}
        table {{ border-collapse: collapse; width: 100%; margin-bottom: 30px; }}
        th, td {{ border: 1px solid #ddd; padding: 8px; text-align: left; }}
        th {{ background-color: #f2f2f2; font-weight: bold; }}
        tr:nth-child(even) {{ background-color: #f9f9f9; }}
        .video {{ background-color: #ffebee; }}
        .audio {{ background-color: #e8f5e8; }}
        .image {{ background-color: #fff3e0; }}
        .file-path {{ font-family: monospace; font-size: 0.9em; }}
    </style>
</head>
<body>
    <div class="header">
        <h1>📊 媒体资源清单</h1>
        <p>扫描时间: {stats['scan_time']}</p>
    </div>

    <div class="stats">
        <div class="stat-card">
            <h3>{stats['total_files']}</h3>
            <p>总文件数</p>
        </div>
        <div class="stat-card">
            <h3>{stats['total_size_mb']} MB</h3>
            <p>总大小</p>
        </div>
        <div class="stat-card">
            <h3>{stats['video_count']}</h3>
            <p>视频文件</p>
        </div>
        <div class="stat-card">
            <h3>{stats['audio_count']}</h3>
            <p>音频文件</p>
        </div>
        <div class="stat-card">
            <h3>{stats['image_count']}</h3>
            <p>图片文件</p>
        </div>
    </div>
"""

        # 为每种媒体类型生成表格
        for media_type, display_name, css_class in [
            ('videos', '🎬 视频文件', 'video'),
            ('audios', '🎵 音频文件', 'audio'),
            ('images', '🖼️ 图片文件', 'image')
        ]:
            files = inventory['files'][media_type]
            if files:
                html_content += f"""
    <h2>{display_name} ({len(files)} 个)</h2>
    <table class="{css_class}">
        <thead>
            <tr>
                <th>文件名</th>
                <th>大小</th>
                <th>扩展名</th>
                <th>修改时间</th>
                <th>文件路径</th>
            </tr>
        </thead>
        <tbody>
"""
                for file_info in files:
                    size_mb = round(file_info['file_size'] / (1024 * 1024), 2)
                    html_content += f"""
            <tr>
                <td>{file_info['file_name']}</td>
                <td>{file_info['file_size']} 字节 ({size_mb} MB)</td>
                <td>{file_info['file_extension']}</td>
                <td>{file_info['modified_time'][:19]}</td>
                <td class="file-path">{file_info['file_path']}</td>
            </tr>
"""
                html_content += """
        </tbody>
    </table>
"""

        html_content += """
</body>
</html>
"""

        with open(output_path, 'w', encoding='utf-8') as f:
            f.write(html_content)

    def _save_markdown(self, inventory: Dict[str, Any], output_path: Path):
        """保存为Markdown格式"""
        stats = inventory['statistics']

        md_content = f"""# 📊 媒体资源清单

**扫描时间**: {stats['scan_time']}

## 📈 统计信息

| 项目 | 数量 |
|------|------|
| 总文件数 | {stats['total_files']} |
| 总大小 | {stats['total_size_mb']} MB |
| 视频文件 | {stats['video_count']} |
| 音频文件 | {stats['audio_count']} |
| 图片文件 | {stats['image_count']} |

## 📁 扩展名统计

| 扩展名 | 文件数 | 总大小(字节) |
|--------|--------|-------------|
"""

        for ext, data in inventory['extensions'].items():
            md_content += f"| {ext} | {data['count']} | {data['size']} |\n"

        # 为每种媒体类型生成表格
        for media_type, display_name in [
            ('videos', '🎬 视频文件'),
            ('audios', '🎵 音频文件'),
            ('images', '🖼️ 图片文件')
        ]:
            files = inventory['files'][media_type]
            if files:
                md_content += f"""
## {display_name} ({len(files)} 个)

| 文件名 | 大小(MB) | 扩展名 | 修改时间 | 文件路径 |
|--------|----------|--------|----------|----------|
"""
                for file_info in files:
                    size_mb = round(file_info['file_size'] / (1024 * 1024), 2)
                    modified_time = file_info['modified_time'][:19]
                    file_path = file_info['file_path'].replace('|', '\\|')  # 转义管道符
                    md_content += f"| {file_info['file_name']} | {size_mb} | {file_info['file_extension']} | {modified_time} | `{file_path}` |\n"

        with open(output_path, 'w', encoding='utf-8') as f:
            f.write(md_content)

    def _save_excel(self, inventory: Dict[str, Any], output_path: Path):
        """保存为Excel格式"""
        try:
            import pandas as pd

            # 创建Excel写入器
            with pd.ExcelWriter(output_path, engine='openpyxl') as writer:

                # 统计信息工作表
                stats = inventory['statistics']
                stats_data = [
                    ['总文件数', stats['total_files']],
                    ['总大小(MB)', stats['total_size_mb']],
                    ['视频文件数', stats['video_count']],
                    ['音频文件数', stats['audio_count']],
                    ['图片文件数', stats['image_count']],
                    ['扫描时间', stats['scan_time']]
                ]
                stats_df = pd.DataFrame(stats_data, columns=['项目', '数值'])
                stats_df.to_excel(writer, sheet_name='统计信息', index=False)

                # 扩展名统计工作表
                ext_data = []
                for ext, data in inventory['extensions'].items():
                    ext_data.append([ext, data['count'], data['size']])
                ext_df = pd.DataFrame(ext_data, columns=['扩展名', '文件数', '总大小(字节)'])
                ext_df.to_excel(writer, sheet_name='扩展名统计', index=False)

                # 为每种媒体类型创建工作表
                for media_type, sheet_name in [
                    ('videos', '视频文件'),
                    ('audios', '音频文件'),
                    ('images', '图片文件')
                ]:
                    files = inventory['files'][media_type]
                    if files:
                        # 准备数据
                        file_data = []
                        for file_info in files:
                            row = {
                                '文件名': file_info['file_name'],
                                '文件路径': file_info['file_path'],
                                '文件大小(字节)': file_info['file_size'],
                                '文件大小(MB)': round(file_info['file_size'] / (1024 * 1024), 2),
                                '扩展名': file_info['file_extension'],
                                'MIME类型': file_info['mime_type'],
                                '创建时间': file_info['created_time'],
                                '修改时间': file_info['modified_time']
                            }

                            # 添加可选字段
                            if file_info.get('file_hash'):
                                row['文件哈希'] = file_info['file_hash']
                            if file_info.get('duration'):
                                row['时长(秒)'] = file_info['duration']
                            if file_info.get('width'):
                                row['宽度'] = file_info['width']
                            if file_info.get('height'):
                                row['高度'] = file_info['height']
                            if file_info.get('fps'):
                                row['帧率'] = file_info['fps']
                            if file_info.get('bitrate'):
                                row['比特率'] = file_info['bitrate']

                            file_data.append(row)

                        # 创建DataFrame并保存
                        df = pd.DataFrame(file_data)
                        df.to_excel(writer, sheet_name=sheet_name, index=False)

                        # 调整列宽
                        worksheet = writer.sheets[sheet_name]
                        for column in worksheet.columns:
                            max_length = 0
                            column_letter = column[0].column_letter
                            for cell in column:
                                try:
                                    if len(str(cell.value)) > max_length:
                                        max_length = len(str(cell.value))
                                except:
                                    pass
                            adjusted_width = min(max_length + 2, 50)
                            worksheet.column_dimensions[column_letter].width = adjusted_width

        except ImportError:
            # 如果没有pandas，使用基本的Excel写入
            self._save_excel_basic(inventory, output_path)

    def _save_excel_basic(self, inventory: Dict[str, Any], output_path: Path):
        """基本Excel保存（不依赖pandas）"""
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill

            wb = Workbook()

            # 删除默认工作表
            wb.remove(wb.active)

            # 统计信息工作表
            stats_ws = wb.create_sheet("统计信息")
            stats = inventory['statistics']

            # 添加标题
            stats_ws['A1'] = '项目'
            stats_ws['B1'] = '数值'
            stats_ws['A1'].font = Font(bold=True)
            stats_ws['B1'].font = Font(bold=True)

            # 添加数据
            stats_data = [
                ['总文件数', stats['total_files']],
                ['总大小(MB)', stats['total_size_mb']],
                ['视频文件数', stats['video_count']],
                ['音频文件数', stats['audio_count']],
                ['图片文件数', stats['image_count']],
                ['扫描时间', stats['scan_time']]
            ]

            for i, (key, value) in enumerate(stats_data, 2):
                stats_ws[f'A{i}'] = key
                stats_ws[f'B{i}'] = value

            # 为每种媒体类型创建工作表
            for media_type, sheet_name in [
                ('videos', '视频文件'),
                ('audios', '音频文件'),
                ('images', '图片文件')
            ]:
                files = inventory['files'][media_type]
                if files:
                    ws = wb.create_sheet(sheet_name)

                    # 添加标题行
                    headers = ['文件名', '文件路径', '文件大小(字节)', '文件大小(MB)', '扩展名', '修改时间']
                    for col, header in enumerate(headers, 1):
                        cell = ws.cell(row=1, column=col, value=header)
                        cell.font = Font(bold=True)
                        cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")

                    # 添加数据
                    for row, file_info in enumerate(files, 2):
                        ws.cell(row=row, column=1, value=file_info['file_name'])
                        ws.cell(row=row, column=2, value=file_info['file_path'])
                        ws.cell(row=row, column=3, value=file_info['file_size'])
                        ws.cell(row=row, column=4, value=round(file_info['file_size'] / (1024 * 1024), 2))
                        ws.cell(row=row, column=5, value=file_info['file_extension'])
                        ws.cell(row=row, column=6, value=file_info['modified_time'][:19])

            wb.save(output_path)

        except ImportError:
            raise ImportError("需要安装 openpyxl 库来支持Excel格式: pip install openpyxl")


def scan_media_resources(directory: Union[str, Path],
                        output_path: Optional[Union[str, Path]] = None,
                        output_format: str = 'json',
                        recursive: bool = True,
                        include_hash: bool = False,
                        include_metadata: bool = True) -> Dict[str, Any]:
    """
    扫描媒体资源的便捷函数

    Args:
        directory: 要扫描的目录
        output_path: 输出文件路径（可选）
        output_format: 输出格式 ('json', 'csv', 'html', 'markdown', 'excel')
        recursive: 是否递归扫描子目录
        include_hash: 是否计算文件哈希值
        include_metadata: 是否提取媒体元数据

    Returns:
        资源清单字典
    """
    scanner = MediaScanner(include_hash=include_hash, include_metadata=include_metadata)
    
    # 扫描目录
    media_files = scanner.scan_directory(directory, recursive=recursive)
    
    # 生成清单
    inventory = scanner.generate_inventory(media_files)
    
    # 保存到文件
    if output_path:
        scanner.save_inventory(inventory, output_path, output_format)
    else:
        # 根据格式确定默认文件名
        format_extensions = {
            'json': '.json',
            'csv': '.csv',
            'html': '.html',
            'markdown': '.md',
            'excel': '.xlsx'
        }
        ext = format_extensions.get(output_format.lower(), '.json')
        default_output = Path(directory) / f"media_inventory{ext}"
        scanner.save_inventory(inventory, default_output, output_format)
    
    return inventory


def main():
    """主函数 - 命令行使用"""
    import argparse
    
    parser = argparse.ArgumentParser(description="媒体资源扫描工具")
    parser.add_argument("directory", help="要扫描的目录路径")
    parser.add_argument("-o", "--output", help="输出文件路径")
    parser.add_argument("-f", "--format", choices=['json', 'csv', 'html', 'markdown', 'excel'],
                       default='json', help="输出格式 (默认: json)")
    parser.add_argument("--no-recursive", action="store_true", help="不递归扫描子目录")
    parser.add_argument("--include-hash", action="store_true", help="计算文件哈希值")
    parser.add_argument("--no-metadata", action="store_true", help="不提取媒体元数据")
    
    args = parser.parse_args()
    
    try:
        inventory = scan_media_resources(
            directory=args.directory,
            output_path=args.output,
            output_format=args.format,
            recursive=not args.no_recursive,
            include_hash=args.include_hash,
            include_metadata=not args.no_metadata
        )
        
        # 显示统计信息
        stats = inventory['statistics']
        print(f"\n📊 扫描统计:")
        print(f"   总文件数: {stats['total_files']}")
        print(f"   总大小: {stats['total_size_mb']} MB")
        print(f"   视频文件: {stats['video_count']}")
        print(f"   音频文件: {stats['audio_count']}")
        print(f"   图片文件: {stats['image_count']}")
        
    except Exception as e:
        print(f"扫描失败: {e}")
        return 1
    
    return 0


if __name__ == "__main__":
    exit(main())
